"""Google Drive service."""

import io
from pathlib import Path

from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload

from .base import BaseGoogleService

# Google Workspace MIME types → export format mapping
_GOOGLE_EXPORT_MAP = {
    "application/vnd.google-apps.document": (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".docx",
    ),
    "application/vnd.google-apps.spreadsheet": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xlsx",
    ),
    "application/vnd.google-apps.presentation": (
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        ".pptx",
    ),
    "application/vnd.google-apps.drawing": ("application/pdf", ".pdf"),
}


class DriveService(BaseGoogleService):
    """Service for Google Drive operations."""

    def __init__(self, drive_service):
        """Initialize Drive service.

        Args:
            drive_service: Google Drive API service
        """
        self.drive = drive_service

    def copy(self, file_id: str, new_title: str = None, folder_id: str = None) -> dict:
        """Copy/duplicate a file.

        Args:
            file_id: Google Drive file ID to copy
            new_title: Title for the copy (default: "Copy of <original>")
            folder_id: Destination folder ID (default: same as original)

        Returns:
            Response with new file info
        """
        try:
            # Get original file info
            original = (
                self.drive.files().get(fileId=file_id, fields="name,mimeType").execute()
            )

            body = {}
            if new_title:
                body["name"] = new_title
            if folder_id:
                body["parents"] = [folder_id]

            copied = (
                self.drive.files()
                .copy(fileId=file_id, body=body, fields="id,name,mimeType,webViewLink")
                .execute()
            )

            return self._format_response(
                data={
                    "originalId": file_id,
                    "originalName": original.get("name"),
                    "newId": copied.get("id"),
                    "newName": copied.get("name"),
                    "type": self._get_file_type(copied.get("mimeType", "")),
                    "url": copied.get("webViewLink"),
                }
            )
        except HttpError as e:
            return self._format_error(self._handle_api_error(e))

    def move(self, file_id: str, folder_id: str) -> dict:
        """Move a file to a different folder.

        Args:
            file_id: Google Drive file ID to move
            folder_id: Destination folder ID

        Returns:
            Response with moved file info
        """
        try:
            # Get current parents
            file_info = (
                self.drive.files().get(fileId=file_id, fields="name,parents").execute()
            )

            previous_parents = ",".join(file_info.get("parents", []))

            # Move file
            updated = (
                self.drive.files()
                .update(
                    fileId=file_id,
                    addParents=folder_id,
                    removeParents=previous_parents,
                    fields="id,name,webViewLink",
                )
                .execute()
            )

            return self._format_response(
                data={
                    "fileId": file_id,
                    "fileName": updated.get("name"),
                    "newFolderId": folder_id,
                    "url": updated.get("webViewLink"),
                }
            )
        except HttpError as e:
            return self._format_error(self._handle_api_error(e))

    def rename(self, file_id: str, new_name: str) -> dict:
        """Rename a file.

        Args:
            file_id: Google Drive file ID
            new_name: New name for the file

        Returns:
            Response with renamed file info
        """
        try:
            updated = (
                self.drive.files()
                .update(
                    fileId=file_id,
                    body={"name": new_name},
                    fields="id,name,webViewLink",
                )
                .execute()
            )

            return self._format_response(
                data={
                    "fileId": file_id,
                    "newName": updated.get("name"),
                    "url": updated.get("webViewLink"),
                }
            )
        except HttpError as e:
            return self._format_error(self._handle_api_error(e))

    def create_folder(self, name: str, parent_id: str = None) -> dict:
        """Create a new folder.

        Args:
            name: Folder name
            parent_id: Parent folder ID (optional, default: root)

        Returns:
            Response with new folder info
        """
        try:
            metadata = {
                "name": name,
                "mimeType": "application/vnd.google-apps.folder",
            }
            if parent_id:
                metadata["parents"] = [parent_id]

            folder = (
                self.drive.files()
                .create(body=metadata, fields="id,name,webViewLink")
                .execute()
            )

            return self._format_response(
                data={
                    "folderId": folder.get("id"),
                    "name": folder.get("name"),
                    "url": folder.get("webViewLink"),
                }
            )
        except HttpError as e:
            return self._format_error(self._handle_api_error(e))

    def delete(self, file_id: str) -> dict:
        """Delete a file from Google Drive.

        Args:
            file_id: Google Drive file ID to delete

        Returns:
            Response confirming deletion
        """
        try:
            # Get file info before deleting
            file_info = (
                self.drive.files().get(fileId=file_id, fields="name,mimeType").execute()
            )
            file_name = file_info.get("name", "Unknown")

            self.drive.files().delete(fileId=file_id).execute()

            return self._format_response(
                data={
                    "fileId": file_id,
                    "fileName": file_name,
                    "deleted": True,
                }
            )
        except HttpError as e:
            return self._format_error(self._handle_api_error(e))

    def share(self, file_id: str, email: str, role: str = "reader") -> dict:
        """Share a file with a user.

        Args:
            file_id: Google Drive file ID
            email: Email address to share with
            role: Permission role (reader, writer, commenter)

        Returns:
            Response with permission info
        """
        valid_roles = ["reader", "writer", "commenter"]
        if role.lower() not in valid_roles:
            return self._format_response(
                success=False,
                error=f"Invalid role: {role}. Use one of: {', '.join(valid_roles)}",
            )

        try:
            permission = (
                self.drive.permissions()
                .create(
                    fileId=file_id,
                    body={"type": "user", "role": role.lower(), "emailAddress": email},
                    sendNotificationEmail=True,
                )
                .execute()
            )

            file_metadata = (
                self.drive.files()
                .get(fileId=file_id, fields="name,webViewLink")
                .execute()
            )

            return self._format_response(
                data={
                    "fileId": file_id,
                    "fileName": file_metadata.get("name"),
                    "sharedWith": email,
                    "role": role.lower(),
                    "permissionId": permission.get("id"),
                    "url": file_metadata.get("webViewLink"),
                }
            )
        except HttpError as e:
            return self._format_error(self._handle_api_error(e))

    def list(self, query: str = None, max_results: int = 20) -> dict:
        """List files in Google Drive.

        Args:
            query: Search query (Google Drive query syntax)
            max_results: Maximum number of results

        Returns:
            Response with file list
        """
        try:
            params = {
                "pageSize": min(max_results, 100),
                "fields": "files(id,name,mimeType,webViewLink,createdTime,modifiedTime)",
                "orderBy": "modifiedTime desc",
            }

            if query:
                params["q"] = query
            else:
                params["q"] = (
                    "mimeType='application/vnd.google-apps.document' or "
                    "mimeType='application/vnd.google-apps.spreadsheet' or "
                    "mimeType='application/vnd.google-apps.presentation'"
                )

            result = self.drive.files().list(**params).execute()

            files = []
            for f in result.get("files", []):
                file_type = self._get_file_type(f.get("mimeType", ""))
                files.append(
                    {
                        "id": f.get("id"),
                        "name": f.get("name"),
                        "type": file_type,
                        "mimeType": f.get("mimeType"),
                        "url": f.get("webViewLink"),
                        "created": f.get("createdTime"),
                        "modified": f.get("modifiedTime"),
                    }
                )

            return self._format_response(
                data={
                    "count": len(files),
                    "files": files,
                }
            )
        except HttpError as e:
            return self._format_error(self._handle_api_error(e))

    def _get_file_type(self, mime_type: str) -> str:
        """Convert MIME type to friendly name.

        Args:
            mime_type: Google Drive MIME type

        Returns:
            Friendly file type name
        """
        type_map = {
            "application/vnd.google-apps.document": "doc",
            "application/vnd.google-apps.spreadsheet": "sheet",
            "application/vnd.google-apps.presentation": "slides",
            "application/vnd.google-apps.folder": "folder",
        }
        return type_map.get(mime_type, "file")

    def upload(
        self,
        file_path: str,
        name: str = None,
        folder_id: str = None,
        share: bool = False,
    ) -> dict:
        """Upload a local file to Google Drive.

        Args:
            file_path: Absolute path to the local file
            name: Filename on Drive (default: original filename)
            folder_id: Destination folder ID (default: root)
            share: Make accessible via link (anyone with link = viewer)

        Returns:
            Response with uploaded file info
        """
        import mimetypes
        from pathlib import Path

        from googleapiclient.http import MediaFileUpload

        path = Path(file_path)
        if not path.exists():
            return self._format_response(
                success=False, error=f"File not found: {file_path}"
            )
        if not path.is_file():
            return self._format_response(
                success=False, error=f"Not a file: {file_path}"
            )

        file_name = name or path.name
        mime_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"

        try:
            metadata = {"name": file_name}
            if folder_id:
                metadata["parents"] = [folder_id]

            media = MediaFileUpload(str(path), mimetype=mime_type, resumable=True)
            result = (
                self.drive.files()
                .create(
                    body=metadata, media_body=media, fields="id,name,webViewLink,size"
                )
                .execute()
            )

            file_id = result.get("id")

            if share:
                self.drive.permissions().create(
                    fileId=file_id,
                    body={"type": "anyone", "role": "reader"},
                ).execute()

            return self._format_response(
                data={
                    "fileId": file_id,
                    "name": result.get("name"),
                    "url": result.get("webViewLink"),
                    "size": result.get("size"),
                    "mimeType": mime_type,
                    "shared": share,
                }
            )
        except HttpError as e:
            return self._format_error(self._handle_api_error(e))

    def read_spreadsheet(
        self,
        file_id: str,
        sheet: str | None = None,
        max_rows: int = 500,
    ) -> dict:
        """Download and read a spreadsheet, returning structured data.

        Supports Google Sheets (exported to xlsx) and uploaded .xlsx files.

        Args:
            file_id: Google Drive file ID
            sheet: Sheet name to read (default: first sheet)
            max_rows: Maximum rows to return (default 500, cap 2000)

        Returns:
            Response with headers, rows, and metadata
        """
        import openpyxl

        max_rows = min(max_rows, 2000)

        try:
            meta = (
                self.drive.files()
                .get(fileId=file_id, fields="id,name,mimeType")
                .execute()
            )
            name = meta.get("name", file_id)
            mime = meta.get("mimeType", "")

            buf = io.BytesIO()

            if mime == "application/vnd.google-apps.spreadsheet":
                export_mime = (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                request = self.drive.files().export_media(
                    fileId=file_id, mimeType=export_mime
                )
            else:
                request = self.drive.files().get_media(fileId=file_id)

            downloader = MediaIoBaseDownload(buf, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()

            buf.seek(0)
            wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)

            sheet_names = wb.sheetnames
            ws = wb[sheet] if sheet and sheet in sheet_names else wb.active

            rows = []
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                str_row = [str(c) if c is not None else "" for c in row]
                if i == 0:
                    headers = str_row
                    continue
                if i > max_rows:
                    break
                rows.append(str_row)

            total_rows = ws.max_row - 1 if ws.max_row else 0
            wb.close()

            return self._format_response(
                data={
                    "fileId": file_id,
                    "fileName": name,
                    "sheetName": ws.title,
                    "allSheets": sheet_names,
                    "headers": headers,
                    "rows": rows,
                    "rowCount": len(rows),
                    "totalRows": total_rows,
                    "truncated": total_rows > max_rows,
                }
            )
        except HttpError as e:
            return self._format_error(self._handle_api_error(e))
        except Exception as e:
            return self._format_response(
                success=False,
                error=f"Failed to read spreadsheet: {e}",
            )

    def download(self, file_id: str, dest_dir: str) -> dict:
        """Download a file from Google Drive to a local directory.

        For Google-native files (Docs, Sheets, Slides) exports to
        Office format (docx, xlsx, pptx). For binary files (PDF, XLSX,
        images, etc.) downloads the original content.

        Args:
            file_id: Google Drive file ID
            dest_dir: Local directory to save the file

        Returns:
            Response with local file path and metadata
        """
        dest = Path(dest_dir)
        dest.mkdir(parents=True, exist_ok=True)

        try:
            meta = (
                self.drive.files()
                .get(fileId=file_id, fields="id,name,mimeType,size")
                .execute()
            )
            name = meta.get("name", file_id)
            mime = meta.get("mimeType", "")

            buf = io.BytesIO()

            if mime in _GOOGLE_EXPORT_MAP:
                # Google-native file → export to Office format
                export_mime, ext = _GOOGLE_EXPORT_MAP[mime]
                request = self.drive.files().export_media(
                    fileId=file_id, mimeType=export_mime
                )
                # Strip any existing extension and add the export one
                stem = Path(name).stem
                name = f"{stem}{ext}"
            else:
                # Binary file → download as-is
                request = self.drive.files().get_media(fileId=file_id)

            downloader = MediaIoBaseDownload(buf, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()

            local_path = dest / name
            local_path.write_bytes(buf.getvalue())

            return self._format_response(
                data={
                    "fileId": file_id,
                    "name": name,
                    "localPath": str(local_path),
                    "mimeType": mime,
                    "size": len(buf.getvalue()),
                }
            )
        except HttpError as e:
            return self._format_error(self._handle_api_error(e))
